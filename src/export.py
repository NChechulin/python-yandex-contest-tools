import csv
from datetime import datetime
from pathlib import Path
from typing import List

from openpyxl import Workbook

from config import Config, Task
from students import Student, TaskResult


class BaseExporter:
    """Base class for Exporter"""

    extension = "notafile"
    students: List[Student]
    tasks: List[Task]
    solved_symbol: str
    banned_symbol: str
    not_solved_symbol: str
    filename = None

    def __init__(self, students: List[Student], config: Config):
        self.students = students
        self.__change_students_names()
        self.students.sort(key=lambda stud: stud.name)

        self.solved_symbol = config.solved_symbol
        self.banned_symbol = config.banned_symbol
        self.not_solved_symbol = config.not_solved_symbol

        self.tasks = config.tasks
        self.tasks.sort(key=lambda task: (len(task.name), task.name))
        self.__set_filename(config.output_dir)

    def __change_students_names(self):
        """Replaces `ё` with `е` in each name to sort names properly"""
        for student in self.students:
            student.name = student.name.replace("ё", "е")

    def _task_result_to_symbol(self, task_result: TaskResult) -> str:
        """Returns a symbol corresponding to student's result"""
        if task_result == TaskResult.Solved:
            return self.solved_symbol
        elif task_result == TaskResult.Banned:
            return self.banned_symbol
        return self.not_solved_symbol

    def __set_filename(self, output_dir: Path) -> str:
        """Returns a filename (stem) where the data will be saved"""
        name = datetime.now().strftime("%Y-%m-%d_%H_%M_%S")
        self.filename = output_dir / f"RESULTS_{name}.{self.extension}"

    def __write_header_to_file(self):
        """Writes column names"""
        raise NotImplementedError()

    def __write_students_data(self):
        """Writes the results of students into the file"""
        raise NotImplementedError()

    def write(self):
        """Writes the all of the data into a file"""
        raise NotImplementedError()


class CSVExporter(BaseExporter):
    extension = "csv"

    def __write_header_to_file(self):
        """Writes column names"""
        columns = ["name"]
        for task in self.tasks:
            columns.append(task.name)

        with open(self.filename, "x") as fh:
            writer = csv.writer(fh)
            writer.writerow(columns)

    def __write_students_data(self):
        """Writes the results of students into the file"""
        with open(self.filename, "a") as fh:
            writer = csv.writer(fh)

            for student in self.students:
                row = [student.name]
                for task in self.tasks:
                    symbol = self._task_result_to_symbol(student.results[task.name])
                    row.append(symbol)

                writer.writerow(row)

    def write(self):
        """Writes the all of the data into a file"""
        self.__write_header_to_file()
        self.__write_students_data()
        print(f"File saved as {self.filename}")


class XLSXExporter(BaseExporter):
    extension = "xlsx"
    workbook: Workbook = None
    worksheet = None

    def __write_header_to_file(self):
        """Writes column names"""
        columns = ["name"]
        for task in self.tasks:
            columns.append(task.name)

        ROW = 1
        for col in range(len(columns)):
            cell = self.worksheet.cell(
                row=ROW, column=col + 1
            )  # indexation starts at 1
            cell.value = columns[col]

    def __write_students_data(self):
        """Writes the results of students into the file"""

        for row, student in enumerate(self.students, start=2):
            self.worksheet.cell(row=row, column=1).value = student.name

            for col, task in enumerate(self.tasks, start=2):
                symbol = self._task_result_to_symbol(student.results[task.name])
                self.worksheet.cell(row=row, column=col).value = symbol

    def __setup_workbook(self):
        """Creates workbook and worksheet"""
        self.workbook = Workbook()
        self.worksheet = self.workbook.active

    def write(self):
        """Writes the all of the data into a file"""
        self.__setup_workbook()
        self.__write_header_to_file()
        self.__write_students_data()

        self.workbook.save(self.filename)
        print(f"File saved as {self.filename}")
